## @file ghgscenario.py
# Generate Excel file from GHG inventory files used for scenarios.
## @package  scen
# Generate Excel file from GHG inventory files used for scenarios.
import glob
import argparse
import numpy as np
import pandas as pd
import openpyxl
import lukeghg.utility.xlmanip as xlp
import lukeghg.crf.ghginventory as ghginv
import lukeghg.crf.crfxmlconstants as crfc

#Lands_FL classes (forestation)
Lands_FL = ['CL-FL','GL-FL','WLpeat-FL','WLother-FL','SE-FL']
#FL_Lands classes (deforestation)
FL_Lands =['FL-CL','FL-GL','FL-WLpeat','FL-WLflooded','FL-WLother','FL-SE']
#Lands_CL classes
Lands_CL = ['FL-CL','GL-CL','WLpeat-CL','WLother-CL','SE-CL']
#Lands_GL classes
Lands_GL = ['FL-GL','CL-GL','WLpeat-GL','WLother-GL','SE-GL']
#Lands_WLpeat classes
Lands_WLpeat = ['FL-WLpeat','CL-WLpeat','GL-WLpeat']
#Lands_WLflooded classes
Lands_WLflooded = ['FL-WLflooded','CL-WLflooded','GL-WLflooded','SE-WLflooded','OL-WLflooded']
#Lands_WLother
Lands_WLother = ['FL-WLother','CL-WLother','GL-WLother'] 
#Lands_WL classes, all three above
Lands_WL = ['Land-WLpeat','Land-WLflooded','Land-WLother']
#Lands_SE classes
Lands_SE = ['FL-SE','CL-SE','GL-SE','WLpeat-SE','WLother-SE']
#WL-WL class
WL_WL = ['WL-WL(peatextraction)','WLother-WLpeat','WL-WL(flooded)','WL-WL(other)','WLpeat-WLother']
#These three classes will tell emissions from peat production, artificial ponds, wetlands etc.
#wether being remaining or converted.
WLpeat_summary = ['WL-WL(peatextraction)']+Lands_WLpeat
WLflooded_summary = ['WL-WL(flooded)']+Lands_WLflooded
WLother_summary = ['WL-WL(other)']+Lands_WLother
          
#Some colors
summary_color='00FFFF00'
error_color='00FF0000'
white_color='FFFFFF'
grey_color='D4D4D4'

class NoInventoryFiles(Exception):
    pass

def check_data_series(ls,start_year:int,end_year:int):
    """
    Check the right length of the time series
    @param ls The time series
    @param start_year Scenario start year
    @param end_year Scenario end year
    @post The time series `ls`  length equal or shorter that scenario years
    @retval ls The time series
    @retval cut True if time series too long False otherwise
    """
    cut = False
    right_length = end_year-start_year+1
    diff = len(ls)-right_length
    if diff > 0:
        ls = ls[:-diff]
        cut = True
    return (ls,cut)

def convert_to_float(x:str, keys_to_zero:bool=True):
    """
    Convert GHG inventory item to float
    @param x The item to be converted
    @param keys_to_zero: True return notation key to 0, False maintain notation keys
    @param x String representing float or notation key or zero
    @pre x is String that can be converted to float or notation key
    @retval x Float, 0 or notation key
    """
    try:
        return float(x)
    except:
        if keys_to_zero:
            return 0
        else:
            return x

def insert_ghg_file_to_dictionary(d:dict,fname:str,uid_mapping_file:str,keys_to_zero:bool=True):
    """
    @param d The dictionary
    @param fname GHG inventory file
    @param uid_mapping_file UID mappings from  CRFReporter version 3.0 to current 5.x one
    @param keys True Convert notation keyes to zero, False maintain notation keys
    @retval d The dictionary filled with the data from `fname`
    """
    datalss = ghginv.ParseGHGInventoryFile(fname,uid_mapping_file)
    for datals in datalss:
        uid=datals.pop(0)
        datals=[convert_to_float(x,keys_to_zero) for x in datals]
        d[uid] = datals
    return d

def create_ghg_file_dictionary(reg_expr:str,uid_mapping_file:str,keys_to_zero:bool=True):
    """
    @param reg_expr  Regular expression to list files in a directory
    @param uid_mapping_file  UID mappings from  CRFReporter version 3.0 to current one
    @param keys_to_zero True convert notation keys to zero, False maintain notain keys
    @retval d The dictionary containing the time series, UID is the dectinary key
    """
    filels = glob.glob(reg_expr)
    d=dict()
    for fname in filels:
        d=insert_ghg_file_to_dictionary(d,fname,uid_mapping_file,keys_to_zero)
    if not d:
        raise NoInventoryFiles
    return d

def read_uid_matrix_file(excel_file:str,skip_rows=4):
    """
    Read the UID matrix sheet from the template Excel
    @param excel_file The Excel template file
    @param skip_rows Empty rows in the sheet to skip
    @retval df The dataframe representing UIDMatrix sheet
    """
    df = pd.read_excel(excel_file,sheet_name="UIDMatrix",skiprows=skip_rows,
                       header=0,engine='openpyxl')
    return df

def read_scenario_template_file(excel_file:str,skip_rows=2):
    df = pd.read_excel(excel_file,sheet_name="LandUse",skiprows=skip_rows,
                       header=0,engine='openpyxl')
    #Change column names to strings, easier to index and slice
    df.columns = [str(x) for x in list(df.columns)]
    return df

def read_scenario_lulucf_file(excel_file:str):
    df = pd.read_excel(excel_file,sheet_name="LULUCF",engine='openpyxl')
    return df

def land_use_classes(df_uid_matrix,index:int=2):
    """
    Slice column names and return land use classes
    @param df_uid_matrix Dataframe UIDMatrix
    @return Column names of the UIDMatrix
    """  
    columnls = list(df_uid_matrix.columns)
    return columnls[index:]

def land_use_class_uid_df(df,name:str):
    """
    @param df Data frame for UIDMatrix
    @param name Land use name
    @retval df2 Return three columns: Number, stock change/emission type and the corresponding land use UID's
    """
    #print("NAME",name)
    df1=df[[df.columns[0],df.columns[1],name]]
    df2=df1[pd.notna(df1[name])]
    #print(df2)
    return df2

def land_use_class_uid_ls(df,name):
    """
    @param df Dataframe for UIDMatrix
    @param name land use name
    @return The list of land use stock change/emission  UID's
    """
    s = df[name]
    return list(s)

def stock_change_name(df,name,uid):
    """
    @param df Dataframe for UIDMatrix
    @param name Land use name
    @param uid Stock change/emission  UID
    @return Stock change/emission type name
    """
    df1 = df[df[name]==uid]
    s = df1[df1.columns[1]]
    ls = list(s)
    return ls[0]

def stock_change_id_number(df,name,uid):
    """
    @param df Dataframe for UID matrix
    @param name Land use name
    @param uid Stock change/emission type UID
    @return Stock change/emission type identification number 
    """
    df1 = df[df[name]==uid]
    s = df1[df1.columns[0]]
    ls = list(s)
    return ls[0]

def stock_change_name_id_number(df,name,uid):
    """
    @param df Dataframe for UIDMatrix
    @param name Stock/emission name
    @param uid Stock/emission uid number
    @return Tuple (n,id_number) where n is stock/emission type name, id_number is the identification number
    """
    n = stock_change_name(df,name,uid)
    id_number = stock_change_id_number(df,name,uid)
    return (n,id_number)

def select_row_number(df_template,id_number):
    """"
    @param df_template UID template sheet or scenario result template sheet
    @param id_number The identification number of the stock emission type
    @retval number Row number corresponding to id_number
    """
    #First (i.e. 0) column is the column name where id_numbers are
    name = df_template.columns[0]
    df_tmp1 = df_template[df_template[name]==id_number]
    number = df_template[df_template[name]==id_number].index[0]
    return number

def add_data_series(df_scen,new_row_ls,start:str,end:str,row_number:int):
    """
    Add scenario template data series to summary data frame (eventually excel sheet)
    @param df_scen Scenario template sheet
    @param new_row_ls: scenario data series to be added to total
    @param start Start year of the inventory
    @apram end End year of the inventory
    @param row_number: the row where the data series belongs to
    @retval df_scen Dataframe with `new_row_ls` added to total 
    """
    length_row = len(new_row_ls)
    length_series = int(end)-int(start)+1
    if length_row < length_series:
        diff = length_series-length_row
        #print("PADDING",diff)
        padding_ls = [0]*diff
        new_row_ls = new_row_ls+padding_ls
    #print("COLS",type(start),type(end))
    #Slicing dataframe with columns as strings or integers is straightforward
    #but mixing strings and integers will get slicing complicated
    df_scen.loc[row_number,start:end]=df_scen.loc[row_number,start:end]+new_row_ls
    return df_scen

def set_data_series(df_scen,new_row_ls,start:str,end:str,row_number:int):
    """
    Set scenario template data series
    @param df_scen Scenario template sheet
    @param new_row_ls Scenario data series
    @param row_number The row where the data series belongs to
    @retval df_scen Dataframe with the `new_row_ls` data series added
    """
    length_row = len(new_row_ls)
    length_series = int(end)-int(start)+1
    if length_row < length_series:
        diff = length_series-length_row
        #print("PADDING",diff)
        padding_ls = [0]*diff
        new_row_ls = new_row_ls+padding_ls
    #print("COLS",type(start),type(end))
    #Slicing dataframe with columns as strings or integers is straightforward
    #but mixing strings and integers will get slicing complicated
    #print("SET DATA:",len(new_row_ls),new_row_ls)
    df_scen.loc[row_number,start:end]=new_row_ls
    #print("SET DATA DONE", len(new_row_ls))
    return df_scen

def write_subtract_formula(sheet,start_col,ncols,result_row,subtract_row):
    """
    Create new subtract cell formulas in a row by appending cells in a `subtract_row`
    @param sheet The Excel sheet
    @param start_col The start column in a row
    @param ncols Number of columns to be used from `start_col`
    @param result_row The result row number
    @param subtract_row The row number for the `subtract_row`
    @retval sheet The Excel `sheet`  with the new subtract formula in `result_row` 
    """
    for i in range(start_col,ncols):
        cell = openpyxl.cell.cell.Cell(sheet,result_row,i)
        cell_subtract = openpyxl.cell.cell.Cell(sheet,subtract_row,i)
        cell_coordinate = cell.coordinate
        cell_subtract_coordinate = cell_subtract.coordinate
        cell_formula=sheet[cell_coordinate].value
        cell_formula=cell_formula+'-'+cell_subtract_coordinate
        sheet[cell_coordinate]=cell_formula
    return sheet
    
def write_co2sum_formula(sheet,start_col,ncols,result_row,row_number_ls,color,scale=1,sheet_ref=""):
    """
    Create summary row as CO2 from its constituent rows
    @param sheet The Excel sheet
    @param start_col start column of the time series
    @param ncols Number of columns (i.e. the length) in the time series
    @param result_row The row number of the sum row
    @paramt row_number_ls The rows to be added to total.
    @param color Background color of the result row
    @param scale The final unit (e.g. scale ktC,CH4,N2O to MtCO2 eq if needed)
    @param sheet_ref Refer to another sheet (e.g. sheet_ref="FL-FL!") if needed
    @retval sheet The excel `sheet`with the `result_row`
    """
    for i in range(start_col,ncols):
        cell = openpyxl.cell.cell.Cell(sheet,result_row,i)
        cell_coordinate = cell.coordinate
        cell_letter = xlp.cut_cell_number(cell_coordinate)
        cell_ls = []
        for row_number in row_number_ls:
            cell_ls.append(cell_letter+str(row_number))
        for cell_name in cell_ls:
            #Fill missing cells with Red
            if sheet[cell_name].value == "":
                #It seems ordering of parameters matters in PatternFill!!
                sheet[cell_name].fill = openpyxl.styles.PatternFill(start_color=error_color,fill_type="solid")
        #Create the excel SUM formula
        formula = '=SUM('
        for cell_name in cell_ls:
            formula = formula+sheet_ref+cell_name+','
        #replace last ',' with ')
        formula = formula[:len(formula)-1]
        formula = formula+')'+'*'+str(scale)
        #print(formula)
        sheet[cell_coordinate]=formula
        #Fill sum rows with yellow
        sheet[cell_coordinate].fill = openpyxl.styles.PatternFill(start_color=color, end_color=color,fill_type = "solid")
    return sheet

def make_zeros(x,index):
    x[index:]=0.0
    return x

def create_sum_rows(sheet,start_year,end_year):
    #1 Biomass Net change
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,9,[7,8],summary_color,1)
    #2 Biomass Total
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,14,range(9,13+1),summary_color,1)
    #3 Direct N2O emissions from N fertilization
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,19,[17,18],summary_color,1)
    #4 Indirect N20 emissions from managed soils
    #Total
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,41,[39,40],summary_color,1)
    #5 Biomass burning
    #Biomass burning total CO2
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,54,[44,49],summary_color,1)
    #Biomass burning total CH4
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,55,[45,50],summary_color,1)
    #Biomass burning total N2O
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,56,[46,51],summary_color,1)
    #5 HWP
    #1 Sawnwood total
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,62,[60,61],summary_color,1)
    #2 Wood panels
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,65,[63,64],summary_color,1)
    #3 Paper and paperboard
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,68,[66,67],summary_color,1)
    #4 HWP Total
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,69,[62,65,68],summary_color,1)
    return sheet

def create_lulucf_sum_rows(sheet,start_year,end_year):
    #Total LULUCF
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,2,[3,6,9,12,15,18],summary_color,1)
    #Forest land
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,3,[4,5],summary_color,1)
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,4,[129],white_color,1,"'FL-FL'!")
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,5,[129],white_color,1,"'Lands_FL'!")
    #Cropland
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,6,[7,8],summary_color,1)
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,7,[129],white_color,1,"'CL-CL'!")
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,8,[129],white_color,1,"'Lands_CL'!")
    #Grassland
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,9,[10,11],summary_color,1)
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,10,[129],white_color,1,"'GL-GL'!")
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,11,[129],white_color,1,"'Lands_GL'!")
    #Wetlands
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,12,[13,14],summary_color,1)
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,13,[129],white_color,1,"'WL_WL'!")
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,14,[129],white_color,1,"'Lands_WL'!")
    #Settlements
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,15,[16,17],summary_color,1)
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,16,[129],white_color,1,"'SE-SE'!")
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,17,[129],white_color,1,"'Lands_SE'!")
    #HWP from FL-FL sheet
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,18,[127],summary_color,1,"'FL-FL'!")
    #Deforestation
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,19,[129],summary_color,1,"'FL_Lands'!")
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,20,[129],white_color,1,"'FL-CL'!")
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,21,[129],white_color,1,"'FL-GL'!")
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,22,[129],white_color,1,"'FL-WLpeat'!")
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,23,[129],white_color,1,"'FL-WLflooded'!")
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,24,[129],white_color,1,"'FL-WLother'!")
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,25,[129],white_color,1,"'FL-SE'!")
    #Check row
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,26,[20,21,22,23,24,25],white_color,1)
    write_subtract_formula(sheet,5,end_year-start_year+1+5,26,19)
    #Afforestation
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,27,[129],summary_color,1,"'Lands_FL'!")
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,28,[129],white_color,1,"'CL-FL'!")
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,29,[129],white_color,1,"'GL-FL'!")
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,30,[129],white_color,1,"'WLpeat-FL'!")
    #write_co2sum_formula(sheet,5,end_year-start_year+1+5,31,[129],white_color,1,"'WLflooded-FL'!")
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,32,[129],white_color,1,"'WLother-FL'!")
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,33,[129],white_color,1,"'SE-FL'!")
    #Check row
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,34,[28,29,30,31,32,33],white_color,1)
    write_subtract_formula(sheet,5,end_year-start_year+1+5,34,27)
    return sheet

def create_MtCO2eq_rows(sheet,MtCO2eq_start_row,start_year,end_year,ch4co2eq,n2oco2eq):
    """
    ---------- Grand totals MtCO2 eq. ----------------
    Create excel formulas for emissions summary as MtCO2eq. 
    MtCO2eq_start_row: Given the start row the block of MtCO2eq summaries should move correctly as a single block
    start_year,end_year: scenario start and end years 
    ch4co2eq,n2oco2eq: GWP, either AR4 (GHG inventory) or AR5 (scenarios, usually)
    """
    #1 Biomass MtCO2 eq.
    #If more detailed classification above is needed this MtCO2 eq part is rather easily moved downwards
    #as a single block by changing the value of  MtCO2eq_start_row for Biomass Gains to the right row.
    #The work needed is to check and correct the lists of rows for each case to be summed and converted to MtCO2eq 
    #Gains
    #C as CO2 for biomass changes sign (atmosphere removal is a positive thing)
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row,[7],summary_color,-1.0*(crfc.ctoco2)/1000.0)
    #Losses
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+1,[8],summary_color,-1.0*(crfc.ctoco2)/1000.0)
    #Net change
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+2,[MtCO2eq_start_row,MtCO2eq_start_row+1],summary_color,1)
    #Dead wood
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+3,[10],summary_color,-1.0*(crfc.ctoco2)/1000.0)
    #Litter
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+4,[11],summary_color,-1.0*(crfc.ctoco2)/1000.0)
    #Mineral soil
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+5,[12],summary_color,-1.0*(crfc.ctoco2)/1000.0)
    #Organic soil
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+6,[13],summary_color,-1.0*(crfc.ctoco2)/1000.0)
    #Total
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+7,range(MtCO2eq_start_row+2,MtCO2eq_start_row+6+1),
                         summary_color,1)
    #2 Direct N20 emissions from N fertilization
    #Inorganic
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+10,[17],summary_color,n2oco2eq/1000.0)
    #Organic
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+11,[18],summary_color,n2oco2eq/1000.0)
    #Total
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+12,[MtCO2eq_start_row+10,MtCO2eq_start_row+11],
                         summary_color,1)
    #3 Drainage and rewetting
    #3.1 Drained organing N2O
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+15,[22],summary_color,n2oco2eq/1000.0)
    #3.2 Drained organic CH4
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+16,[23],summary_color,ch4co2eq/1000.0)
    #Drained organic total
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+17,[MtCO2eq_start_row+15,MtCO2eq_start_row+16],
                         summary_color,1)
    #3.3 Rewetted organing N2O
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+18,[25],summary_color,n2oco2eq/1000.0)
    #3.4 Rewetted organing CH4
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+19,[26],summary_color,ch4co2eq/1000.0)
    #Rewetted organic total
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+20,[MtCO2eq_start_row+18,MtCO2eq_start_row+19],summary_color,1)
    #3.5 Drained mineral N20
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+21,[28],summary_color,n2oco2eq/1000.0)
    #3.6 Drained mineral CH4
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+22,[29],summary_color,ch4co2eq/1000.0)
    #Drained mineral total
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+23,[MtCO2eq_start_row+21,MtCO2eq_start_row+22],summary_color,1)
    #3.7 Rewetted mineral N20
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+24,[31],summary_color,n2oco2eq/1000.0)
    #3.8 Rewetted mineral CH4
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+25,[32],summary_color,ch4co2eq/1000.0)
    #Rewetted mineral total
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+26,[MtCO2eq_start_row+24,MtCO2eq_start_row+25],summary_color,1)
    #4 Direct N2O emissions mineralization (total)
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+29,[36],summary_color,n2oco2eq/1000.0)
    #5 Indirect N2O emissions managed soils
    #5.1 Atmospheric deposition N2O
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+32,[39],summary_color,n2oco2eq/1000.0)
    #5.2 Nitrogen leaching and run-off N2O
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+33,[40],summary_color,n2oco2eq/1000.0)
    #Indirect N2O managed soils total
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+34,[MtCO2eq_start_row+32,MtCO2eq_start_row+33],summary_color,1)
    #6 Biomass burning
    #Controlled burning CO2
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+37,[44],summary_color,(crfc.ctoco2)/1000.0)
    #Controlled burning CH4
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+38,[45],summary_color,ch4co2eq/1000.0)
    #Controlled burning N2O
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+39,[46],summary_color,n2oco2eq/1000.0)
    #Total
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+40,
                         [MtCO2eq_start_row+37,MtCO2eq_start_row+38,MtCO2eq_start_row+39],summary_color,1)
    #Wildfires CO2
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+42,[49],summary_color,(crfc.ctoco2)/1000.0)
    #Wildfires CH4
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+43,[50],summary_color,ch4co2eq/1000.0)
    #Wildfires N2O
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+44,[51],summary_color,n2oco2eq/1000.0)
    #Total
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+45,
                         [MtCO2eq_start_row+42,MtCO2eq_start_row+43,MtCO2eq_start_row+44],summary_color,1)
    #8 HWP
    #Sawnwood
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+48,[62],summary_color,-1.0*(crfc.ctoco2)/1000.0)
    #Wood panels
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+49,[65],summary_color,-1.0*(crfc.ctoco2)/1000.0)
    #Paper and paper board
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+50,[68],summary_color,-1.0*(crfc.ctoco2)/1000.0)
    #Total
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+51,
                         [MtCO2eq_start_row+48,MtCO2eq_start_row+49,MtCO2eq_start_row+50],summary_color,1)
    #GRAND TOTAL MtCO2eq, HWP Total NOT included 
    write_co2sum_formula(sheet,5,end_year-start_year+1+5,MtCO2eq_start_row+53,
                         [MtCO2eq_start_row+7,MtCO2eq_start_row+12,MtCO2eq_start_row+17,MtCO2eq_start_row+20,
                          MtCO2eq_start_row+23,MtCO2eq_start_row+26,MtCO2eq_start_row+29,MtCO2eq_start_row+34,
                          MtCO2eq_start_row+40,MtCO2eq_start_row+45],summary_color,1)
    return sheet

def create_land_use_summary_formulas(sheet,start_col:int,ncols:int,result_row_ls:list,start_year:int,end_year:int,color,sheet_ref_ls:list):
    """
    sheet: excel_sheet
    start_col: start column for time series
    ncols: the nth column up to which to fill with formulas 
    result_row: result row to contain the summation formula
    start_year: start year of the inventory
    end_year: end year of the inventory
    color: cell fill color
    sheet_ref_ls: reference to sheets to sum up
    """
    #For each row in the list
    for result_row  in result_row_ls:
        #For each column from start_col to ncols-1 (e.g. range(1,5) return [1,2,3,4])
        for i in range(start_col,ncols):
            #Find the cell in the result row and i'th columns
            cell = openpyxl.cell.cell.Cell(sheet,result_row,i)
            #Then cell cooridinates
            cell_coordinate = cell.coordinate
            #Drop the number part
            cell_letter = xlp.cut_cell_number(cell_coordinate)
            #From the list of references to the sheets to sum up
            #create a summation string that contains also right row number
            #We use the same row to collect results and attach the formula.
            s=""
            for sheet_ref in sheet_ref_ls:
                s=s+sheet_ref+cell_letter+str(result_row)+'+'
            #Remove last '+'
            s = s[:-1]
            formula = "="+s
            #Attach the formula to the cell
            sheet[cell_coordinate]=formula
            #Fill sum rows with yellow, otherwise white
            sheet[cell_coordinate].fill = openpyxl.styles.PatternFill(start_color=color, end_color=color,fill_type = "solid")
    return sheet

def create_scenario_excel(scen_excel_file:str,scen_files_reg_expr:str,scen_template_file:str,uid_300_500_file:str,
                          start_year:int,end_year:int,ch4co2eq,n2oco2eq,formulas:bool,gwp_str:str):
    """
    Main program entry. Create scenario excel from scenario inventory files.
    @param scen_excel_file The name of the output excel file
    @param scen_files_reg_expr Regular expression to list scenario inventory files 
    @param scen_template_file The Excel template file to generate `scen_excel_file`
    @param uid_300_500_file The UID conversion file to convert CRFReporter 3.0.0 UIDs to CRFReport 5.x UIDs
    @param start_year Scenario start year (1990)
    @param end_year Scenario end year (2050)
    @param ch4co2eq CH4 to CO2eq Global Warming Potential
    @param n2oco2eq N2O to CO2eq Global Warming Potential
    @param formulas True create Excel formulas in summary sheets, False calculate the values in summary sheets
    @param gwp_str Either GWP4 or GWP5 (default)
    @retval writer The representation of the Excel file `scen_excel_file`
    """
    missing_uid_dict = dict()
    long_time_series_dict = dict()
    writer = pd.ExcelWriter(scen_excel_file,engine='openpyxl')
    #Read inventory to dictionary: UID:time series
    uid_time_series_dictionary = create_ghg_file_dictionary(scen_files_reg_expr,uid_300_500_file,True)
    df_uid = read_uid_matrix_file(scen_template_file)
    #Column values, i.e. land use classes from UIDMatrix
    ls = land_use_classes(df_uid)
    #LULUCF summary sheet
    df_lulucf = read_scenario_lulucf_file(scen_template_file)
    df_scen_template = read_scenario_template_file(scen_template_file)
    #Land-Forestland transition (afforestation)
    df_lfl = df_scen_template.copy()
    #This simply initializes the emission time series to 0 for Land->Forestland
    df_lfl[df_lfl[df_lfl.columns[0]].astype(str).str.isnumeric()]=\
    df_lfl[df_lfl[df_lfl.columns[0]].astype(str).str.isnumeric()].apply(lambda x: make_zeros(x,3),axis=1)
    #df_lfl.iloc[:,3:]=0.0
    #Forestland-Land transition (deforestation)
    df_fll = df_scen_template.copy()
    #This simply initializes the emission time series to 0 for Forestland->Land
    df_fll[df_fll[df_fll.columns[0]].astype(str).str.isnumeric()]=\
    df_fll[df_fll[df_fll.columns[0]].astype(str).str.isnumeric()].apply(lambda x: make_zeros(x,3),axis=1)
    #df_fll.iloc[:,3:]=0.0
    #Land to Cropland
    df_lcl =  df_scen_template.copy()
    df_lcl[df_lcl[df_lcl.columns[0]].astype(str).str.isnumeric()]=\
    df_lcl[df_lcl[df_lcl.columns[0]].astype(str).str.isnumeric()].apply(lambda x: make_zeros(x,3),axis=1)
    #Land to Grassland
    df_lgl =  df_scen_template.copy()
    df_lgl[df_lgl[df_lgl.columns[0]].astype(str).str.isnumeric()]=\
    df_lgl[df_lgl[df_lgl.columns[0]].astype(str).str.isnumeric()].apply(lambda x: make_zeros(x,3),axis=1)
    #Land to Peatland
    df_lwlpeat =  df_scen_template.copy()
    df_lwlpeat[df_lwlpeat[df_lwlpeat.columns[0]].astype(str).str.isnumeric()]=\
    df_lwlpeat[df_lwlpeat[df_lwlpeat.columns[0]].astype(str).str.isnumeric()].apply(lambda x: make_zeros(x,3),axis=1)
    #Land to Flooded
    df_lwlflooded =  df_scen_template.copy()
    df_lwlflooded[df_lwlflooded[df_lwlflooded.columns[0]].astype(str).str.isnumeric()]=\
    df_lwlflooded[df_lwlflooded[df_lwlflooded.columns[0]].astype(str).str.isnumeric()].apply(lambda x: make_zeros(x,3),axis=1)
    #Land to WLother
    df_lwlother =  df_scen_template.copy()
    df_lwlother[df_lwlother[df_lwlother.columns[0]].astype(str).str.isnumeric()]=\
    df_lwlother[df_lwlother[df_lwlother.columns[0]].astype(str).str.isnumeric()].apply(lambda x: make_zeros(x,3),axis=1)
    #Land to WL, all three above
    df_lwl =  df_scen_template.copy()
    df_lwl[df_lwl[df_lwl.columns[0]].astype(str).str.isnumeric()]=\
    df_lwl[df_lwl[df_lwl.columns[0]].astype(str).str.isnumeric()].apply(lambda x: make_zeros(x,3),axis=1)
    #Land to SE
    df_lse =  df_scen_template.copy()
    df_lse[df_lse[df_lse.columns[0]].astype(str).str.isnumeric()]=\
    df_lse[df_lse[df_lse.columns[0]].astype(str).str.isnumeric()].apply(lambda x: make_zeros(x,3),axis=1)
    #WL to WL
    df_wlwl =  df_scen_template.copy()
    df_wlwl[df_wlwl[df_wlwl.columns[0]].astype(str).str.isnumeric()]=\
    df_wlwl[df_wlwl[df_wlwl.columns[0]].astype(str).str.isnumeric()].apply(lambda x: make_zeros(x,3),axis=1)
    #Three new sheets will tell emissions from peat production, artificial ponds, wetlands etc.
    #wether remaining or converted.
    #Wetlands peat
    df_wlpeat_summary =  df_scen_template.copy()
    df_wlpeat_summary[df_wlpeat_summary[df_wlpeat_summary.columns[0]].astype(str).str.isnumeric()]=\
    df_wlpeat_summary[df_wlpeat_summary[df_wlpeat_summary.columns[0]].astype(str).str.isnumeric()].apply(lambda x: make_zeros(x,3),axis=1)
    #Wetlands flooded
    df_wlflooded_summary = df_scen_template.copy()
    df_wlflooded_summary[df_wlflooded_summary[df_wlflooded_summary.columns[0]].astype(str).str.isnumeric()]=\
    df_wlflooded_summary[df_wlflooded_summary[df_wlflooded_summary.columns[0]].astype(str).str.isnumeric()].apply(lambda x: make_zeros(x,3),axis=1)
    #Wetlands other
    df_wlother_summary = df_scen_template.copy()
    df_wlother_summary[df_wlother_summary[df_wlother_summary.columns[0]].astype(str).str.isnumeric()]=\
    df_wlother_summary[df_wlother_summary[df_wlother_summary.columns[0]].astype(str).str.isnumeric()].apply(lambda x: make_zeros(x,3),axis=1)
    for class_name in ls:
        print("LAND USE",class_name)
        #Initialize missing uid list
        missing_uid_dict[class_name]=[np.NaN]
        #Return for three columns from the UIDMatrix: Number, stock change/emission type, class_name
        df_uid_class=land_use_class_uid_df(df_uid,class_name)
        #Make a list of uids in class_name column
        uid_ls = land_use_class_uid_ls(df_uid_class,class_name)
        df_scen_new = df_scen_template.copy()
        #1.Add time series to dataframe
        for uid in uid_ls:
            #Strip whitespace from the beginning and at the end
            uid = uid.strip()
            (name,id_number) = stock_change_name_id_number(df_uid,class_name,uid)
            if uid in uid_time_series_dictionary:
                check_land_class = 0
                #print("UID",uid)
                data_series_ls = uid_time_series_dictionary[uid]
                orig_length = len(data_series_ls)
                (data_series_ls,cut) = check_data_series(data_series_ls,start_year,end_year)
                if orig_length != (end_year - start_year + 1):
                    long_time_series_dict[uid]=orig_length
                row_number = select_row_number(df_scen_template,id_number)
                df_scen_new = set_data_series(df_scen_new,data_series_ls,str(start_year),str(end_year),row_number)
                if class_name in Lands_FL:
                    df_lfl=add_data_series(df_lfl,data_series_ls,str(start_year),str(end_year),row_number)
                    check_land_class = 1
                if class_name in FL_Lands:
                     df_fll=add_data_series(df_fll,data_series_ls,str(start_year),str(end_year),row_number)
                     check_land_class =1
                if class_name in Lands_CL:
                    df_lcl = add_data_series(df_lcl,data_series_ls,str(start_year),str(end_year),row_number)
                    check_land_class = 1
                if class_name in Lands_GL:
                    df_lgl = add_data_series(df_lgl,data_series_ls,str(start_year),str(end_year),row_number)
                    check_land_class = 1
                if class_name in Lands_WLpeat:
                    df_lwlpeat = add_data_series(df_lwlpeat,data_series_ls,str(start_year),str(end_year),row_number)
                    #Land to WL,df_lwl, includes all three cases: peat, flooded and other
                    df_lwl = add_data_series(df_lwl,data_series_ls,str(start_year),str(end_year),row_number)
                    check_land_class = 1
                if class_name in Lands_WLflooded:
                    df_lwlflooded = add_data_series(df_lwlflooded,data_series_ls,str(start_year),str(end_year),row_number)
                    df_lwl = add_data_series(df_lwl,data_series_ls,str(start_year),str(end_year),row_number)
                    check_land_class = 1
                if class_name in Lands_WLother:
                    df_lwlother = add_data_series(df_lwlother,data_series_ls,str(start_year),str(end_year),row_number)
                    df_lwl = add_data_series(df_lwl,data_series_ls,str(start_year),str(end_year),row_number)
                    check_land_class = 1
                if class_name in Lands_SE:
                    df_lse = add_data_series(df_lse,data_series_ls,str(start_year),str(end_year),row_number)
                    check_land_class = 1
                if class_name in WL_WL:
                    #print("WL_WL",class_name)
                    df_wlwl = add_data_series(df_wlwl,data_series_ls,str(start_year),str(end_year),row_number)
                    check_land_class = 1
                if class_name in WLpeat_summary:
                    df_wlpeat_summary = add_data_series(df_wlpeat_summary,data_series_ls,str(start_year),str(end_year),row_number)
                    check_land_class = 1
                if class_name in WLflooded_summary:
                    df_wlflooded_summary = add_data_series(df_wlflooded_summary,data_series_ls,str(start_year),str(end_year),row_number)
                    check_land_class = 1
                if class_name in WLother_summary:
                    df_wlother_summary = add_data_series(df_wlother_summary,data_series_ls,str(start_year),str(end_year),row_number)
                    check_land_class = 1
                if check_land_class == 0:
                    pass
                    #print("Land class not in summary sheets",uid,class_name)
            else:
                print("MISSING",class_name,uid)
                missing_uid_dict[class_name].append(uid)
        #Convert years to numbers for excel (no alerts in excel)
        column_ls = df_scen_new.columns
        year_ls = column_ls[2:]
        column_ls = ["Number","Source","Unit"]+list(range(start_year,start_year+len(year_ls)-1))
        df_scen_new.columns = column_ls
        df_scen_new.to_excel(writer,sheet_name=class_name)
        #2.Add sum rows
        sheets = writer.sheets
        sheet = sheets[class_name]
        sheet = create_sum_rows(sheet,start_year,end_year)
        #3. ---------- Grand totals MtCO2 eq. ----------------
        sheet = create_MtCO2eq_rows(sheet,76,start_year,end_year,ch4co2eq,n2oco2eq)
    #Excel sheet for missing UID values
    missing_uid_df = pd.DataFrame.from_dict(missing_uid_dict,orient='index')
    #missing_uid_df = missing_uid_df.dropna()
    print("Creating information sheets")
    uid_sheet_name="UID not in Inventory"
    missing_uid_df.to_excel(writer,sheet_name=uid_sheet_name)
    #Excel sheet for too long time series
    long_time_series_df = pd.DataFrame.from_dict(long_time_series_dict,orient='index')
    long_series_sheet_name = "Time series length != "+str(end_year-start_year+1)
    long_time_series_df.to_excel(writer,sheet_name=long_series_sheet_name)
    #Excel sheet for GWP
    ls = [['CO2','CH4','N2O'],[1,ch4co2eq,n2oco2eq]]
    df = pd.DataFrame(ls)
    #Mark in sheet name if GWP4 or GWP5 (default)
    gwp_sheet="GWP"+gwp_str
    df.to_excel(writer,sheet_name=gwp_sheet)
    df_lulucf.to_excel(writer,sheet_name='LULUCF')
    column_ls = df_lfl.columns
    year_ls = column_ls[2:]
    column_ls = ["Number","Source","Unit"]+list(range(start_year,start_year+len(year_ls)-1))
    df_fll.columns = column_ls
    df_lfl.columns = column_ls
    df_lcl.columns = column_ls
    df_lgl.columns = column_ls
    df_lse.columns = column_ls
    df_lwl.columns = column_ls
    df_wlwl.columns = column_ls
    df_lwlpeat.columns = column_ls
    df_lwlflooded.columns = column_ls
    df_lwlother.columns = column_ls
    df_wlpeat_summary.columns = column_ls
    df_wlflooded_summary.columns = column_ls
    df_wlother_summary.columns = column_ls
    df_fll.to_excel(writer,sheet_name='FL_Lands')
    df_lfl.to_excel(writer,sheet_name='Lands_FL')
    df_lcl.to_excel(writer,sheet_name='Lands_CL')
    df_lgl.to_excel(writer,sheet_name='Lands_GL')
    df_lse.to_excel(writer,sheet_name='Lands_SE')
    df_lwl.to_excel(writer,sheet_name='Lands_WL')
    df_wlwl.to_excel(writer,sheet_name='WL_WL')
    df_lwlpeat.to_excel(writer,sheet_name='Lands_WLpeat')
    df_lwlflooded.to_excel(writer,sheet_name='Lands_WLflooded')
    df_lwlother.to_excel(writer,sheet_name='Lands_WLother')
    df_wlpeat_summary.to_excel(writer,sheet_name='WLpeat_summary')
    df_wlflooded_summary.to_excel(writer,sheet_name='WLflooded_summary')
    df_wlother_summary.to_excel(writer,sheet_name='WLother_summary')
    sheets = writer.sheets
    sheet_wlpeat_summary = sheets['WLpeat_summary']
    sheet_wlflooded_summary = sheets['WLflooded_summary']
    sheet_wlother_summary = sheets['WLother_summary']
    sheet_fll = sheets['FL_Lands']
    sheet_lfl = sheets['Lands_FL']
    sheet_lcl = sheets['Lands_CL']
    sheet_lgl = sheets['Lands_GL']
    sheet_lwlpeat = sheets['Lands_WLpeat']
    sheet_lwlflooded = sheets['Lands_WLflooded']
    sheet_lwlother = sheets['Lands_WLother']
    sheet_lwl = sheets['Lands_WL']
    sheet_wlwl = sheets['WL_WL']
    sheet_lse = sheets['Lands_SE']
    sheet_lulucf = sheets['LULUCF']
    print("Creating summary sheets")
    #Summation for LULUCF and land change classes: FL_Lands, Lands_FL etc.
    sheet_wlpeat_summary = create_sum_rows(sheet_wlpeat_summary,start_year,end_year)
    sheet_wlpeat_summary = create_MtCO2eq_rows(sheet_wlpeat_summary,76,start_year,end_year,ch4co2eq,n2oco2eq)
    sheet_wlflooded_summary = create_sum_rows(sheet_wlflooded_summary,start_year,end_year)
    sheet_wlflooded_summary = create_MtCO2eq_rows(sheet_wlflooded_summary,76,start_year,end_year,ch4co2eq,n2oco2eq)
    sheet_wlother_summary = create_sum_rows(sheet_wlother_summary,start_year,end_year)
    sheet_wlother_summary = create_MtCO2eq_rows(sheet_wlother_summary,76,start_year,end_year,ch4co2eq,n2oco2eq)
    sheet_fll = create_sum_rows(sheet_fll,start_year,end_year)
    sheet_fll = create_MtCO2eq_rows(sheet_fll,76,start_year,end_year,ch4co2eq,n2oco2eq)
    sheet_lfl = create_sum_rows(sheet_lfl,start_year,end_year)
    sheet_lfl = create_MtCO2eq_rows(sheet_lfl,76,start_year,end_year,ch4co2eq,n2oco2eq)
    sheet_lcl = create_sum_rows(sheet_lcl,start_year,end_year)
    sheet_lcl = create_MtCO2eq_rows(sheet_lcl,76,start_year,end_year,ch4co2eq,n2oco2eq)
    sheet_lgl = create_sum_rows(sheet_lgl,start_year,end_year)
    sheet_lgl = create_MtCO2eq_rows(sheet_lgl,76,start_year,end_year,ch4co2eq,n2oco2eq)
    sheet_lwlpeat = create_sum_rows(sheet_lwlpeat,start_year,end_year)
    sheet_lwlpeat = create_MtCO2eq_rows(sheet_lwlpeat,76,start_year,end_year,ch4co2eq,n2oco2eq)
    sheet_lwlflooded = create_sum_rows(sheet_lwlflooded,start_year,end_year)
    sheet_lwlflooded = create_MtCO2eq_rows(sheet_lwlflooded,76,start_year,end_year,ch4co2eq,n2oco2eq)
    sheet_lwlother = create_sum_rows(sheet_lwlother,start_year,end_year)
    sheet_lwlother = create_MtCO2eq_rows(sheet_lwlother,76,start_year,end_year,ch4co2eq,n2oco2eq)
    sheet_lwl = create_sum_rows(sheet_lwl,start_year,end_year)
    sheet_lwl = create_MtCO2eq_rows(sheet_lwl,76,start_year,end_year,ch4co2eq,n2oco2eq)
    sheet_wlwl = create_sum_rows(sheet_wlwl,start_year,end_year)
    sheet_wlwl = create_MtCO2eq_rows(sheet_wlwl,76,start_year,end_year,ch4co2eq,n2oco2eq)
    sheet_lse = create_sum_rows(sheet_lse,start_year,end_year)
    sheet_lse = create_MtCO2eq_rows(sheet_lse,76,start_year,end_year,ch4co2eq,n2oco2eq)
    sheet_lulucf = create_lulucf_sum_rows(sheet_lulucf,start_year,end_year)
    if formulas:
        print("Creating formulas in summary sheets")
        create_land_use_summary_formulas(sheet_fll,5,end_year-start_year+1+5,list(range(7,130)),start_year,end_year,grey_color,
                                         ["'FL-CL'!","'FL-GL'!","'FL-WLpeat'!","'FL-WLflooded'!","'FL-WLother'!","'FL-SE'!"])
        create_land_use_summary_formulas(sheet_lfl,5,end_year-start_year+1+5,list(range(7,130)),start_year,end_year,grey_color,
                                         ["'CL-FL'!","'GL-FL'!","'WLpeat-FL'!","'WLother-FL'!","'SE-FL'!"])
        create_land_use_summary_formulas(sheet_lcl,5,end_year-start_year+1+5,list(range(7,130)),start_year,end_year,grey_color,
                                         ["'FL-CL'!","'GL-CL'!","'WLpeat-CL'!","'WLother-CL'!","'SE-CL'!"])
        create_land_use_summary_formulas(sheet_lgl,5,end_year-start_year+1+5,list(range(7,130)),start_year,end_year,grey_color,
                                         ["'FL-GL'!","'CL-GL'!","'WLPeat-GL'!","'WLother-GL'!","'SE-GL'!"])
        create_land_use_summary_formulas(sheet_lse,5,end_year-start_year+1+5,list(range(7,130)),start_year,end_year,grey_color,
                                         ["'FL-SE'!","'CL-SE'!","'GL-SE'!","'WLpeat-SE'!","'WLother-SE'!"])
        create_land_use_summary_formulas(sheet_lwl,5,end_year-start_year+1+5,list(range(7,130)),start_year,end_year,grey_color,
                                         ["Lands_WLpeat!","Lands_WLflooded!","Lands_WLother!"])
        create_land_use_summary_formulas(sheet_wlwl,5,end_year-start_year+1+5,list(range(7,130)),start_year,end_year,grey_color,
                                         ["'WL-WL(peatextraction)'!","'WLother-WLpeat'!","'WL-WL(flooded)'!","'WL-WL(other)'!","'WLpeat-WLother'!"])
        create_land_use_summary_formulas(sheet_lwlpeat,5,end_year-start_year+1+5,list(range(7,130)),start_year,end_year,grey_color,
                                         ["'FL-WLpeat'!","'CL-WLpeat'!","'GL-WLpeat'!"])
        create_land_use_summary_formulas(sheet_lwlflooded,5,end_year-start_year+1+5,list(range(7,130)),start_year,end_year,grey_color,
                                         ["'FL-WLflooded'!","'CL-WLflooded'!","'GL-WLflooded'!","'SE-WLflooded'!","'OL-WLflooded'!"])
        create_land_use_summary_formulas(sheet_lwlother,5,end_year-start_year+1+5,list(range(7,130)),start_year,end_year,grey_color,
                                         ["'FL-WLother'!","'CL-WLother'!","'GL-WLother'!"])
        create_land_use_summary_formulas(sheet_wlother_summary,5,end_year-start_year+1+5,list(range(7,130)),start_year,end_year,grey_color,
                                         ["'WL-WL(other)'!","'FL-WLother'!","'CL-WLother'!","'GL-WLother'!"])
        create_land_use_summary_formulas(sheet_wlflooded_summary,5,end_year-start_year+1+5,list(range(7,130)),start_year,end_year,grey_color,
                                         ["'WL-WL(flooded)'!","'FL-WLflooded'!","'CL-WLflooded'!","'GL-WLflooded'!","'SE-WLflooded'!","'OL-WLflooded'!"])
        create_land_use_summary_formulas(sheet_wlpeat_summary,5,end_year-start_year+1+5,list(range(7,130)),start_year,end_year,grey_color,
                                         ["'WL-WL(peatextraction)'!","'FL-WLpeat'!","'CL-WLpeat'!","'GL-WLpeat'!"])
    #Rotate sheets from the end to the beginning
    #ad hoc numbers to get summary sheets first
    print("Rotating sheets: summary sheets first")
    workbook = writer.book
    workbook.move_sheet('WLpeat_summary',-112)
    workbook.move_sheet('WLflooded_summary',-112)
    workbook.move_sheet('WLother_summary',-112)
    workbook.move_sheet('Lands_WLother',-112)
    workbook.move_sheet('Lands_WLflooded',-112)
    workbook.move_sheet('Lands_WLpeat',-112)
    workbook.move_sheet('WL_WL',-112)
    workbook.move_sheet('Lands_WL',-112)
    workbook.move_sheet('Lands_SE',-112)
    workbook.move_sheet('Lands_GL',-112)
    workbook.move_sheet('Lands_CL',-112)
    workbook.move_sheet('Lands_FL',-112)
    workbook.move_sheet('FL_Lands',-112)
    workbook.move_sheet('LULUCF',-112)
    workbook.move_sheet(gwp_sheet,-112)
    workbook.move_sheet(long_series_sheet_name,-112)
    workbook.move_sheet(uid_sheet_name,-112)
    workbook.active=0
    return writer

