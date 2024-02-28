import os
import glob
from xml.etree.ElementTree import ElementTree as ET
from xml.etree.ElementTree import Element,SubElement
from optparse import OptionParser as OP
import pandas as pd
import openpyxl 
from lukeghg.crf.uid340to500mapping import MapUID340to500, Create340to500UIDMapping
from lukeghg.crf.crfxmlfunctions import ConvertFloat
## @file
# Compare two inventories.
# Compare two inventories for susceptible values, deficiences and errors.
# These include  too large deviations between inventroy values, changes in notiation keys
# and missing inventory data. Comparison is made between two time series with the same UIDs.

##  Color for titles
title_color='00FFFF00'
##  Color for errors
error_color='00FF0000'
## Color for notation key changes
nk_change_color ='00FF9900'
## Color for change from notation key to number  
nk_change_to_number_color='0099CC00'
## Notation keys in CRFReporter
nkls = ['IE','NE','NO','NA', 'R']
## Methods in  CRFReporter
methodls = ['T1','T2','T3']

## @defgroup Dictionary Create dictionary for time series.
# Dictionary for time series.
# @{
def CreateDictionary(dirfilels:list):
    """! Create a dictionary that associate a UID (dictinary key) to its time series,comment and file name.
    Format is {UID1:(time_series1,comment1,file1),UID2:(time_series2,comment2,file2),....}
    @param dirfilels: List of directory files
    @return  dict: Dictionary \sa dict.
    """
    ## @var $dict
    #  Return value, dictionary of UIDs keys and times series, comments and file name as data
    dict={}
    for file in dirfilels:
        #print(file)
        f=open(file)
        datals = [x.rpartition('#') for x in f.readlines() if x.count('#') != 1]
        f.close()
        for tuple3 in datals:
            comment = tuple3[0]
            comment=comment.lstrip('#')
            time_series = tuple3[2].split()
            if len(time_series)==0:
                print("Found empty line:",file)
            else:
                uid = time_series.pop(0)
                uid_new = uid.strip('{}')
                #print(uid_new,time_series,len(time_series))
                dict[uid_new]=(time_series,comment,file)
    return dict
## @}

## @defgroup SignGroup Sign of a number.
# Determine the sign of a number. \sa Sign \sa SameSign
# @{
def Sign(x):
    """! Sign of `x`, return -1 if `x` < 0, otherwise 1.
    @param x: number
    @return -1 or 1
    
    """
    if x < 0:
        return -1
    else:
        return 1

def SameSign(x,y):
    """! Return True if same Sign, otherwise return False.
    @param x: number
    @param y: number
    @return Sign(x)==Sign(y)
    \sa Sign
    """
    s1 = Sign(x)
    s2 = Sign(y)
    return s1 == s2
## @}

## @defgroup CmpInventory Compare two inventories
# Compare two inventories and write differences to Excel file.
# The comparison of two inventory years proceeds in four steps
# in `CompareTwoInventoryYears` documented in submodules:
#  + Check inventory: First check if differences exist
#  + Dictionaries for differences: If differences found crete corresponfing dictionaries
#  + Excel result: Create data frames for Excel sheet (from difference dictionaries) and write Excel file 
#    + Coloring cells: Define colors for Excel sheet cells
#
# \sa CompareTwoInventoryYears
# @{
def CompareTwoInventoryYears(dict1:dict,dict2:dict,tolerance,uidnotincurrentyear:dict,uidnotinpreviousyear:dict,collect_NO:bool,file_out:str):
    """! Comapare two inventory years collected into the two dictionaries.
    Each time series has UID identifier that is the key to the dictionary.
    The value returned by the UID key is a three tuple (time_series,comment, file).
    This way using time series comments and files names a more informative output 
    can be generated for possibly flawed results.
    @param dict1: The current inventory year.
    @param dict2: The previous inventory year (or some else).
    @param tolerance: The difference (%) or greater that is not accepted between two inventpry values.
    @param uidnotincurrentyear:  UIDs not in current year
    @param uidnotinpreviousyear:  UIDs not in previous year
    @param collect_NO: collect NO notation keys to an excel sheet
    @param file_out: output file for possible errata in GHG inventory .
    @return None: Write an Excel output file to `file_out`. 
    
    The Excel file has six sections: 
       1. Values above tolerance.
       2. Changes in notation keys.
       3. Changes in methods. 
       4. Zeroes (number 0) found (not allowed in GHG inventory).  
       5. UIDs not in current year. \sa uidnotincurrentyear.
       6. UIDs not in previous year. \sa uidnotinpreviousyear.
    """
## @}
##
    current_year_keyls=dict1.keys()
    ## @defgroup DiffDict Dictionaries for differences.
    # @ingroup CmpInventory
    # @{
    # Collect differences into dictionaries.
    # Collect differences into three dictionaries and print out to Excel file_out.
    # Each three dictionary contains the UID as the key and time_series, the comment and the file
    # as 3-tuple data.
    # @var $method_diff_dict
    # Differences in methods
    method_diff_dict={}
    # @var $nk_diff_dict
    # Differences in notation keys
    nk_diff_dict={}
    # @var $result_nk_diff
    # Differences in results
    result_diff_dict={}
    ## @var $zero_numbers_dict
    # Zeroes (number 0) found in inventory.
    zero_numbers_dict={}
    ## @}
    for key in current_year_keyls:
        ##  @defgroup DiffSet Check inventory differences.
        # @ingroup CmpInventory
        # Check if inventory differences.
        # Check differences between two time series year by year for methods, notation keys, changes in results and for zeroes.
        # If one or more found the respective set will have True. Write  the two time series compared to output file.
        # @{
        nk_diff_set=set()
        method_diff_set=set()
        value_diff_set=set()
        zero_numbers_set=set()
        ## @var $relative_change_ls
        # List containing relative changes between two time series. \sa CompareTwoInventoryYears and the parameter `tolerance`.
        relative_change_ls=[]
        ## @} 
        if key in dict2:
            (time_series_current,comment_current,file_current) = dict1[key]
            (time_series_prev,comment_prev,file_prev) = dict2[key]
            #start comparing values, zip drops the last value from the current time series
            len_diff = len(time_series_current) - len(time_series_prev)
            #KP might have whole time series calculated or not
            #If not pad with '-' marks
            for item in range(0,len_diff-1):
                time_series_prev.insert(0,'-')
            for (val_current,val_prev) in zip(time_series_current,time_series_prev):
                #Start checking the reason for differences
                if (val_current != val_prev and val_prev != '-'):
                    stringls1=val_current.split(',')
                    stringls2=val_prev.split(',')
                    #Either of the value is a notation key
                    if (stringls1[0] in nkls) or (stringls2[0] in nkls):
                        nk_diff_set.add(True)
                    elif (stringls1[0] in methodls) or (stringls2[0] in methodls):
                        method_diff_set.add(True)
                    #Both must be numbers
                    #Check if nominator is zero
                    else:
                        f1 = float(val_current)
                        f2 = float(val_prev)
                        #Nominator might be zero
                        try:
                            relative_change_percentage=((f1-f2)/abs(f2))*100.0
                            if relative_change_percentage >= tolerance:
                                value_diff_set.add(True)
                        except ZeroDivisionError:
                            zero_numbers_set.add(True)
                #Special case: both might have zero value
                try:
                    f1 = float(val_current)
                    f2 = float(val_prev)
                    if f1 == 0.0 or f2 == 0.0:
                        zero_numbers_set.add(True)
                except ValueError:
                    pass
            #Time series is checked collect the results
            if True in nk_diff_set:
                nk_diff_dict[key]=(time_series_current,comment_current,time_series_prev,file_current,file_prev)
            if True in method_diff_set:
                method_diff_dict[key]=(time_series_current,comment_current,time_series_prev,file_current,file_prev)
            if True in zero_numbers_set:
                zero_numbers_dict[key]=(time_series_current,comment_current,time_series_prev,file_current,file_prev)
            #Calculate relative change for each year
            if True in value_diff_set:
                #More work: calculate difference for each year
                for (f1,f2) in zip(time_series_current,time_series_prev):
                    try:
                        relative_change=((float(f1)-float(f2))/abs(float(f2)))*100.0
                        relative_change_ls.append(relative_change)
                    except ValueError:
                        #Not possible to calculate relative change 
                        relative_change_ls.append('-')
                    except ZeroDivisionError:
                        relative_change_ls.append('ZERO')
                #Collect the relative changes
                result_diff_dict[key]=(time_series_current,comment_current,time_series_prev,relative_change_ls,file_current,file_prev)
    ## @defgroup ExcelFile Excel result file.
    # @ingroup CmpInventory
    # Write the results for each type of differences to Excel file.
    # @{
    result_diff_keyls=result_diff_dict.keys()
    data_frame_list=[]
    data_frame_list.append(['Section Relative change more than '+str(tolerance)+'%'])
    for key in result_diff_keyls:
        (time_series_current,comment,time_series_prev,relative_change_ls,file_current,file_prev)=result_diff_dict[key]
        len_diff = len(time_series_current)-len(time_series_prev)
        for item in range(0,len_diff-1):
            time_series_prev.insert(0,'-')
        for item in relative_change_ls:
            try:
                item=float(format(float(item),'.3f'))
            except ValueError:
                item=item
        time_series_prev=[ConvertFloat(x) for x in time_series_prev]
        time_series_current=[ConvertFloat(x) for x in time_series_current]
        relative_change_ls=[ConvertFloat(x) for x in relative_change_ls]
        data_frame_list.append([key,comment,file_prev]+time_series_prev)
        data_frame_list.append([key,comment,file_current]+time_series_current)
        data_frame_list.append([key,"","Difference %"]+relative_change_ls)
    data_frame_list.append(['------'])
    data_frame_list.append(['Section Change in Notation key'])
    result_nk_diff_keyls = nk_diff_dict.keys()
    for key in result_nk_diff_keyls:
        (time_series_current,comment,time_series_prev,file_current,file_prev)=nk_diff_dict[key]
        len_diff = len(time_series_current)-len(time_series_prev)
        for item in range(0,len_diff-1):
            time_series_prev.insert(0,'-')
        time_series_prev=[ConvertFloat(x) for x in time_series_prev]
        time_series_current=[ConvertFloat(x) for x in time_series_current]
        data_frame_list.append([key,comment,file_prev]+time_series_prev)
        data_frame_list.append([key,comment,file_current]+time_series_current)
        data_frame_list.append([key,"","Notation key change"]+time_series_current)
    data_frame_list.append(['------'])
    data_frame_list.append(['Section Change in Methods'])
    result_method_diff_keyls = method_diff_dict.keys()
    for key in result_method_diff_keyls:
        (time_series_current,comment,time_series_prev,file_current,file_prev)=method_diff_dict[key]
        len_diff = len(time_series_current)-len(time_series_prev)
        for item in range(0,len_diff-1):
            time_series_prev.insert(0,'-')
        data_frame_list.append([key,comment,file_prev]+time_series_prev)
        data_frame_list.append([key,comment,file_current]+time_series_current)
    data_frame_list.append(['------'])
    data_frame_list.append(['Section Found zeros (i.e. number 0)'])
    result_zero_numbers_keyls = zero_numbers_dict.keys()
    for key in result_zero_numbers_keyls:
        (time_series_current,comment,time_series_prev,file_current,file_prev)=zero_numbers_dict[key]
        len_diff = len(time_series_current)-len(time_series_prev)
        for item in range(0,len_diff-1):
            time_series_prev.insert(0,'-')
        time_series_prev=[ConvertFloat(x) for x in time_series_prev]
        time_series_current=[ConvertFloat(x) for x in time_series_current]
        data_frame_list.append([key,comment,file_prev]+time_series_prev)
        data_frame_list.append([key,comment,file_current]+time_series_current)
    data_frame_list.append(['UID not in current year','Comment','File'])
    for uid in uidnotincurrentyear:
        (time_series_prev,comment_prev,file_prev) = dict2[uid]
        data_frame_list.append([uid,comment_prev,file_prev])
    data_frame_list.append(['UID not in previous year','Comment','File'])
    for uid in uidnotinpreviousyear:
        (time_series_current,comment_current,file_current) = dict1[uid]
        data_frame_list.append([uid,comment_current,file_current])
    data_frame=pd.DataFrame(data_frame_list)
    writer = pd.ExcelWriter(file_out,engine='openpyxl')
    data_frame.to_excel(writer,sheet_name="Inventory Check")
    sheets = writer.sheets
    sheet_active = sheets["Inventory Check"]
    ## @}
    #
    ## @defgroup Coloring Coloring cells
    # @ingroup ExcelFile
    # Coloring susceptible values.
    # Coloring red values that are above tolerance
    # Coloring orange notation key changes. Coloring red when number (=result) changed to notation key.
    # Coloring yellow the beginning of sections
    # @{
    for row in sheet_active.iter_rows(min_col=4,max_col=4):
        for cell in  row:
            if type(cell.value) is str and "Diff" in cell.value:
                #Collecting cells from a single row starting column 5
                for col in sheet_active.iter_rows(min_row=cell.row,max_row=cell.row, min_col=5):
                    for col_cell in col:
                        #Check if value >=tolerance
                        if type(col_cell.value) is float and col_cell.value >= tolerance:
                            col_cell.fill =  openpyxl.styles.PatternFill(start_color=error_color, end_color=error_color,fill_type = "solid")
    #Coloring orange notation key changes, coloring red when number (=result) changed to notation key
    for row in sheet_active.iter_rows(min_col=4,max_col=4):
        for cell in row:
            if type(cell.value) is str and "Notation key" in cell.value:
                #Collecting cells from a single row starting column 5
                for col in sheet_active.iter_rows(min_row=cell.row,max_row=cell.row, min_col=5):
                    #Compare with a original value (two rows above, same column)
                    for col_cell in col:
                        cell_orig = sheet_active.cell(row=col_cell.row-2,column=col_cell.column)
                        #Change in notation key
                        if (type(col_cell.value)==str and type(cell_orig.value) == str) and (col_cell.value != cell_orig.value) and (cell_orig.value !="") :
                            col_cell.fill =  openpyxl.styles.PatternFill(start_color=nk_change_color, end_color=nk_change_color,fill_type = "solid")
                        #Calculated value to Notation key
                        if (type(col_cell.value)==str and type(cell_orig.value) == float) and (cell_orig.value !="") :
                            col_cell.fill =  openpyxl.styles.PatternFill(start_color=error_color, end_color=error_color,fill_type = "solid")
                        #Notation key to calculated value
                        if (type(col_cell.value)==float and type(cell_orig.value) == str) and (cell_orig.value !="") :
                            col_cell.fill =  openpyxl.styles.PatternFill(start_color=nk_change_to_number_color, end_color=nk_change_to_number_color,fill_type = "solid")
    #Coloring yellow the beginning of sections
    for row in sheet_active.iter_rows(min_col=2,max_col=2):
        for cell in row:
            #Note paranthesis to denote either Section or UID in cell value
            if type(cell.value) is str and ("Section" in cell.value or "UID" in cell.value):
                cell.fill = openpyxl.styles.PatternFill(start_color=title_color, end_color=title_color,fill_type = "solid")
    #Collect NO notation keys
    if collect_NO == True:
        df_NO_ls = []
        print("Collecting NO notation keys") 
        current_year_keyls=dict1.keys()
        for key in current_year_keyls:
            (time_series,comment,file_name) = dict1[key]
            if 'NO' in time_series:
                df_NO_ls.append([key,comment,file_name]+time_series)
        df = pd.DataFrame(df_NO_ls)
        df.to_excel(writer,sheet_name="NO Notation keys")
    writer.save()
    ## @}

